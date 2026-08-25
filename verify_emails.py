import argparse
import json
import os
import random
import time
from pathlib import Path

import requests


def disable_broken_local_proxy():
    proxy_names = [
        "HTTP_PROXY",
        "HTTPS_PROXY",
        "ALL_PROXY",
        "http_proxy",
        "https_proxy",
        "all_proxy",
    ]

    for name in proxy_names:
        value = os.environ.get(name, "")
        if "127.0.0.1:9" in value or "localhost:9" in value:
            os.environ.pop(name, None)


disable_broken_local_proxy()


# ============================================================
# DEFAULT CONFIGURATION
# ============================================================

DEFAULT_SPREADSHEET_URL = ""
DEFAULT_WORKSHEET_NAME = ""

# Replit-friendly: either upload credentials.json or paste the whole JSON into
# a Replit Secret named GOOGLE_CREDENTIALS_JSON.
DEFAULT_CREDENTIALS_FILE = "credentials.json"
GOOGLE_CREDENTIALS_ENV = "GOOGLE_CREDENTIALS_JSON"

# Your current layout:
# C = email input, W = Valid/Invalid, X = score/reason.
DEFAULT_EMAIL_COLUMN = "C"
DEFAULT_RESULT_COLUMN = "W"
DEFAULT_SCORE_COLUMN = "X"

API_URL = "https://rapid-email-verifier.fly.dev/api/validate/batch"
REQUEST_TIMEOUT = 120
CHECKPOINT_FILE = "email_verify_checkpoint.json"

# Slow defaults to avoid hammering the verifier server.
DEFAULT_BATCH_SIZE = 20
DEFAULT_BATCH_DELAY = 15
DEFAULT_RETRY_DELAY = 30
DEFAULT_JITTER = 5


def column_to_index(column):
    if isinstance(column, int):
        return column

    text = str(column).strip().upper()
    if not text:
        raise ValueError("Column cannot be blank.")
    if text.isdigit():
        return int(text)

    value = 0
    for char in text:
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Invalid column: {column}")
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def index_to_column(index):
    index = int(index)
    if index < 1:
        raise ValueError("Column index must be 1 or greater.")

    letters = []
    while index:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def clean_email(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def sleep_slow(seconds, jitter):
    delay = seconds + random.uniform(0, jitter)
    print(f"Sleeping {delay:.1f}s before continuing...")
    time.sleep(delay)


def load_checkpoint(path):
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return set(int(row) for row in data.get("completed_rows", []))
    except Exception:
        return set()


def save_checkpoint(path, completed_rows):
    data = {
        "completed_rows": sorted(completed_rows),
        "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=2)


def get_reason(result):
    status = str(result.get("status", "")).upper()
    validations = result.get("validations", {}) or {}

    if status == "INVALID_FORMAT":
        return "Invalid email format"
    if status == "INVALID_DOMAIN":
        return "Invalid / nonexistent domain"
    if status == "DISPOSABLE":
        return "Disposable email"

    if validations.get("syntax") is False:
        return "Invalid email format"
    if validations.get("domain_exists") is False:
        return "Domain does not exist"
    if validations.get("mx_records") is False:
        return "No valid MX records"
    if validations.get("is_disposable") is True:
        return "Disposable email"
    if validations.get("is_role_based") is True:
        return "Role-based email"

    if status:
        return status.replace("_", " ").title()
    return "Verification failed"


def calculate_score(result):
    status = str(result.get("status", "")).upper()
    validations = result.get("validations", {}) or {}

    if status == "VALID":
        return 85 if validations.get("is_role_based") is True else 100
    if status == "PROBABLY_VALID":
        return 85
    if status in {"INVALID_FORMAT", "INVALID_DOMAIN", "DISPOSABLE"}:
        return 0
    if validations.get("syntax") is False:
        return 0
    if validations.get("domain_exists") is False:
        return 0
    if validations.get("mx_records") is False:
        return 0
    return 0


def convert_result(result):
    status = str(result.get("status", "")).upper()
    score = calculate_score(result)

    if status in {"VALID", "PROBABLY_VALID"}:
        if status == "PROBABLY_VALID":
            return "Valid", f"{score} - {get_reason(result)}"
        if result.get("validations", {}).get("is_role_based") is True:
            return "Valid", f"{score} - Role-based email"
        return "Valid", str(score)

    return "Invalid", f"{score} - {get_reason(result)}"


def verify_batch(emails, retry_delay, jitter):
    payload = {"emails": emails}

    for attempt in range(1, 4):
        try:
            response = requests.post(API_URL, json=payload, timeout=REQUEST_TIMEOUT)

            if response.status_code == 429 or 500 <= response.status_code <= 599:
                raise RuntimeError(
                    f"Server returned HTTP {response.status_code}: {response.text[:300]}"
                )

            response.raise_for_status()
            data = response.json()
            results = data.get("results")
            if not isinstance(results, list):
                raise RuntimeError(f"Unexpected API response: {data}")
            return results

        except Exception as exc:
            print(f"API error - attempt {attempt}/3: {exc}")
            if attempt < 3:
                sleep_slow(retry_delay * attempt, jitter)

    raise RuntimeError("Batch failed after 3 attempts.")


class GoogleSheetSource:
    def __init__(self, spreadsheet_url, worksheet_name, credentials_file):
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        credentials_json = os.getenv(GOOGLE_CREDENTIALS_ENV)
        if credentials_json:
            credentials = Credentials.from_service_account_info(
                json.loads(credentials_json),
                scopes=scopes,
            )
        else:
            credentials = Credentials.from_service_account_file(
                credentials_file,
                scopes=scopes,
            )

        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_url(spreadsheet_url)
        self.worksheet = spreadsheet.worksheet(worksheet_name)
        self.name = f"Google Sheet worksheet '{worksheet_name}'"

    def read_rows(self):
        return self.worksheet.get_all_values()

    def write_results(self, row_updates, result_col, score_col):
        result_letter = index_to_column(result_col)
        score_letter = index_to_column(score_col)
        updates = []

        for row_number, verification, score_reason in row_updates:
            updates.append(
                {
                    "range": f"{result_letter}{row_number}:{score_letter}{row_number}",
                    "values": [[verification, score_reason]],
                }
            )

        self.worksheet.batch_update(
            updates,
            value_input_option="USER_ENTERED",
        )

    def save(self):
        return


class ExcelSource:
    def __init__(self, excel_file, worksheet_name):
        from openpyxl import load_workbook

        self.path = Path(excel_file)
        self.workbook = load_workbook(self.path)
        self.worksheet = self.workbook[worksheet_name] if worksheet_name else self.workbook.active
        self.name = f"Excel file '{self.path}' sheet '{self.worksheet.title}'"

    def read_rows(self):
        rows = []
        for row in self.worksheet.iter_rows(values_only=True):
            rows.append(["" if value is None else str(value) for value in row])
        return rows

    def write_results(self, row_updates, result_col, score_col):
        for row_number, verification, score_reason in row_updates:
            self.worksheet.cell(row=row_number, column=result_col).value = verification
            self.worksheet.cell(row=row_number, column=score_col).value = score_reason

    def save(self):
        self.workbook.save(self.path)


def build_source(args):
    if args.source == "sheets":
        return GoogleSheetSource(
            args.spreadsheet_url,
            args.worksheet or DEFAULT_WORKSHEET_NAME,
            args.credentials_file,
        )
    return ExcelSource(args.excel_file, args.worksheet)


def ask_row_range():
    while True:
        try:
            start_row = int(input("\nStart Row: ").strip())
            end_row = int(input("End Row: ").strip())

            if start_row < 1:
                print("Start row must be 1 or greater.")
                continue
            if end_row < start_row:
                print("End row must be greater than or equal to Start Row.")
                continue
            return start_row, end_row
        except ValueError:
            print("Please enter valid row numbers.")


def collect_pending(rows, start_row, end_row, email_col, result_col, completed_rows):
    pending = []
    skipped = 0
    blank = 0
    checkpoint_skipped = 0

    for row_number in range(start_row, end_row + 1):
        row = rows[row_number - 1]
        email = clean_email(row[email_col - 1]) if len(row) >= email_col else ""
        existing_result = str(row[result_col - 1]).strip() if len(row) >= result_col else ""

        if row_number in completed_rows:
            checkpoint_skipped += 1
            continue
        if not email:
            blank += 1
            continue
        if existing_result:
            skipped += 1
            continue

        pending.append({"row": row_number, "email": email})

    return pending, skipped, blank, checkpoint_skipped


def run(args):
    email_col = column_to_index(args.email_column)
    result_col = column_to_index(args.result_column)
    score_col = column_to_index(args.score_column)

    print("=" * 70)
    print("EMAIL VERIFICATION - SLOW SAFE MODE")
    print("=" * 70)
    print(f"Email column: {index_to_column(email_col)}")
    print(f"Result column: {index_to_column(result_col)}")
    print(f"Score/reason column: {index_to_column(score_col)}")
    print(f"Batch size: {args.batch_size}")
    print(f"Delay between batches: {args.batch_delay}s + jitter")

    if args.batch_size < 1 or args.batch_size > 100:
        raise ValueError("Batch size must be between 1 and 100.")

    if args.start_row and args.end_row:
        start_row, end_row = args.start_row, args.end_row
    else:
        start_row, end_row = ask_row_range()

    print("\nConnecting/opening source...")
    source = build_source(args)
    print(f"Connected: {source.name}")

    print("\nReading rows...")
    rows = source.read_rows()

    if not rows:
        print("No rows found.")
        return
    if start_row > len(rows):
        print(f"Start row {start_row} is beyond the last row ({len(rows)}).")
        return

    end_row = min(end_row, len(rows))
    completed_rows = load_checkpoint(args.checkpoint)

    pending, skipped, blank, checkpoint_skipped = collect_pending(
        rows,
        start_row,
        end_row,
        email_col,
        result_col,
        completed_rows,
    )

    print(f"\nRows selected: {start_row} to {end_row}")
    print(f"Emails to verify: {len(pending)}")
    print(f"Already completed in result column: {skipped}")
    print(f"Already completed in checkpoint: {checkpoint_skipped}")
    print(f"Blank emails: {blank}")

    if not pending:
        print("\nNothing to verify.")
        return

    email_to_rows = {}
    for item in pending:
        email_to_rows.setdefault(item["email"], []).append(item["row"])

    unique_emails = list(email_to_rows.keys())
    total_batches = (len(unique_emails) + args.batch_size - 1) // args.batch_size
    processed_unique = 0

    print(f"Unique emails: {len(unique_emails)}")

    for batch_start in range(0, len(unique_emails), args.batch_size):
        batch_number = batch_start // args.batch_size + 1
        batch_emails = unique_emails[batch_start:batch_start + args.batch_size]

        print(
            f"\nBatch {batch_number}/{total_batches} "
            f"({len(batch_emails)} unique emails)"
        )

        try:
            results = verify_batch(batch_emails, args.retry_delay, args.jitter)
        except Exception as exc:
            print(f"Batch failed and was skipped: {exc}")
            sleep_slow(args.retry_delay, args.jitter)
            continue

        row_updates = []
        for result in results:
            email = clean_email(result.get("email", ""))
            if not email:
                continue

            verification, score_reason = convert_result(result)
            for row_number in email_to_rows.get(email, []):
                row_updates.append((row_number, verification, score_reason))
                print(f"Row {row_number}: {email} -> {verification} | {score_reason}")

        if row_updates:
            source.write_results(row_updates, result_col, score_col)
            for row_number, _, _ in row_updates:
                completed_rows.add(row_number)

        source.save()
        save_checkpoint(args.checkpoint, completed_rows)

        processed_unique += len(batch_emails)
        print(f"Saved {len(row_updates)} row updates.")
        print(f"Progress: {processed_unique}/{len(unique_emails)} unique emails processed.")

        if batch_start + args.batch_size < len(unique_emails):
            sleep_slow(args.batch_delay, args.jitter)

    print("\n" + "=" * 70)
    print("VERIFICATION COMPLETED")
    print("=" * 70)
    print(f"Column {index_to_column(result_col)} = Valid / Invalid")
    print(f"Column {index_to_column(score_col)} = Score / Reason")


def parse_args():
    parser = argparse.ArgumentParser(description="Slow email verification for Google Sheets or Excel.")
    parser.add_argument("--source", choices=["sheets", "excel"], default="sheets")

    parser.add_argument("--spreadsheet-url", default=DEFAULT_SPREADSHEET_URL)
    parser.add_argument("--worksheet")
    parser.add_argument("--credentials-file", default=DEFAULT_CREDENTIALS_FILE)
    parser.add_argument("--excel-file", default="emails.xlsx")

    parser.add_argument("--email-column", default=DEFAULT_EMAIL_COLUMN)
    parser.add_argument("--result-column", default=DEFAULT_RESULT_COLUMN)
    parser.add_argument("--score-column", default=DEFAULT_SCORE_COLUMN)

    parser.add_argument("--start-row", type=int)
    parser.add_argument("--end-row", type=int)
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--batch-delay", type=int, default=DEFAULT_BATCH_DELAY)
    parser.add_argument("--retry-delay", type=int, default=DEFAULT_RETRY_DELAY)
    parser.add_argument("--jitter", type=int, default=DEFAULT_JITTER)
    parser.add_argument("--checkpoint", default=CHECKPOINT_FILE)

    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
