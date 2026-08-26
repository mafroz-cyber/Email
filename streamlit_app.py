import json
import os
import re
import tempfile
import time
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st


def disable_broken_local_proxy():
    for name in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"):
        value = os.environ.get(name, "")
        if "127.0.0.1:9" in value or "localhost:9" in value:
            os.environ.pop(name, None)


disable_broken_local_proxy()

from verify_emails import (
    DEFAULT_EMAIL_COLUMN,
    DEFAULT_RESULT_COLUMN,
    DEFAULT_SCORE_COLUMN,
    DEFAULT_SPREADSHEET_URL,
    GOOGLE_CREDENTIALS_ENV,
    ExcelSource,
    GoogleSheetSource,
    clean_email,
    collect_pending,
    column_to_index,
    convert_result,
    index_to_column,
    verify_batch,
)


APP_DIR = Path(".email_verify_app")
SAVED_CREDENTIALS_FILE = APP_DIR / "google_credentials.json"
SAVED_CONFIG_FILE = APP_DIR / "config.json"

st.set_page_config(page_title="EmailVerify Pro", page_icon="@", layout="wide")


def inject_css():
    st.markdown(
        """
<style>
.block-container {
  padding-top: 1.5rem;
  padding-bottom: 3rem;
  max-width: 1220px;
}
.ev-title {
  font-size: 2.1rem;
  font-weight: 800;
  margin: 0 0 .25rem 0;
  background: linear-gradient(90deg, #2a78d6, #6f42e8);
  -webkit-background-clip: text;
  background-clip: text;
  -webkit-text-fill-color: transparent;
  display: inline-block;
}
.ev-subtitle {
  opacity: .7;
  margin-bottom: 1.4rem;
}
.ev-band {
  border: 1px solid rgba(128, 128, 128, .25);
  border-radius: 10px;
  padding: 1rem 1.1rem;
  margin: .75rem 0;
}
.ev-kicker {
  opacity: .65;
  font-size: .8rem;
  text-transform: uppercase;
  letter-spacing: .06em;
  font-weight: 750;
  margin-bottom: .5rem;
}
.ev-step {
  border-left: 3px solid var(--primary-color, #2563eb);
  padding-left: .9rem;
  margin: .75rem 0;
}
div[data-testid="stMetric"] {
  border: 1px solid rgba(128, 128, 128, .25);
  border-radius: 10px;
  padding: .9rem;
  transition: border-color .15s ease;
}
div[data-testid="stMetric"]:hover {
  border-color: rgba(42, 120, 214, .55);
}
div[data-testid="stVerticalBlockBorderWrapper"] {
  border-radius: 12px !important;
}
div[data-testid="stButton"] > button {
  border-radius: 8px;
  font-weight: 650;
}
div[data-testid="stButton"] > button[kind="primary"] {
  box-shadow: 0 2px 10px rgba(42, 120, 214, .35);
}
hr {
  margin: 1.1rem 0;
}
</style>
""",
        unsafe_allow_html=True,
    )


def app_header(title, subtitle):
    st.markdown(f'<div class="ev-title">{title}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="ev-subtitle">{subtitle}</div>', unsafe_allow_html=True)


def load_saved_config():
    try:
        with open(SAVED_CONFIG_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}


def save_config(config):
    APP_DIR.mkdir(exist_ok=True)
    with open(SAVED_CONFIG_FILE, "w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2)


def save_credentials(credentials_text):
    credentials = json.loads(credentials_text)
    APP_DIR.mkdir(exist_ok=True)
    with open(SAVED_CREDENTIALS_FILE, "w", encoding="utf-8") as handle:
        json.dump(credentials, handle, indent=2)
    return credentials.get("client_email", "")


def load_credentials_text():
    if SAVED_CREDENTIALS_FILE.exists():
        return SAVED_CREDENTIALS_FILE.read_text(encoding="utf-8")

    try:
        secret_value = st.secrets.get(GOOGLE_CREDENTIALS_ENV, "")
        if secret_value:
            return str(secret_value)
    except Exception:
        pass

    return ""


def remove_saved_credentials():
    if SAVED_CREDENTIALS_FILE.exists():
        SAVED_CREDENTIALS_FILE.unlink()


def get_json_client_email(credentials_text):
    if not credentials_text.strip():
        return ""
    try:
        return json.loads(credentials_text).get("client_email", "")
    except Exception:
        return ""


def unique_email_map(pending):
    email_to_rows = {}
    for item in pending:
        email_to_rows.setdefault(item["email"], []).append(item["row"])
    return email_to_rows


def parse_pasted_emails(text):
    if not text:
        return []

    seen = set()
    emails = []
    for token in re.split(r"[\s,;]+", text.strip()):
        email = clean_email(token)
        if email and email not in seen:
            seen.add(email)
            emails.append(email)
    return emails


def build_excel_source(uploaded_file, worksheet_name):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    return ExcelSource(tmp_path, worksheet_name), tmp_path


@st.cache_resource(show_spinner="Connecting to Google Sheet...")
def cached_google_worksheet(spreadsheet_url, worksheet_name, credentials_text):
    if credentials_text.strip():
        json.loads(credentials_text)
        os.environ[GOOGLE_CREDENTIALS_ENV] = credentials_text.strip()
    return GoogleSheetSource(spreadsheet_url, worksheet_name, "credentials.json")


@st.cache_data(ttl=60, show_spinner="Reading sheet rows...")
def cached_google_rows(spreadsheet_url, worksheet_name, credentials_text, cache_bust):
    source = cached_google_worksheet(spreadsheet_url, worksheet_name, credentials_text)
    return source.read_rows()


@st.cache_resource(show_spinner="Opening spreadsheet...")
def cached_google_spreadsheet(spreadsheet_url, credentials_text):
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    if credentials_text.strip():
        credentials = Credentials.from_service_account_info(
            json.loads(credentials_text), scopes=scopes
        )
    else:
        credentials = Credentials.from_service_account_file("credentials.json", scopes=scopes)

    client = gspread.authorize(credentials)
    return client.open_by_url(spreadsheet_url)


@st.cache_data(ttl=120, show_spinner="Fetching sheet tabs...")
def cached_worksheet_titles(spreadsheet_url, credentials_text):
    spreadsheet = cached_google_spreadsheet(spreadsheet_url, credentials_text)
    return [worksheet.title for worksheet in spreadsheet.worksheets()]


def read_excel_rows_from_bytes(file_bytes, sheet_name):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])
    return rows


def make_source(config):
    if config["mode"] == "Google Sheets":
        credentials_text = config.get("credentials_text", "") or load_credentials_text()
        return cached_google_worksheet(
            config["spreadsheet_url"],
            config["worksheet_name"],
            credentials_text.strip(),
        ), None

    uploaded_file = st.session_state.get("excel_file")
    if uploaded_file is None:
        raise ValueError("Upload an Excel file in Setup first.")
    return build_excel_source(uploaded_file, config.get("excel_sheet", ""))


def read_rows_for(source, config):
    if config["mode"] == "Google Sheets":
        credentials_text = (config.get("credentials_text", "") or load_credentials_text()).strip()
        cache_bust = st.session_state.get("sheet_cache_bust", 0)
        return cached_google_rows(
            config["spreadsheet_url"], config["worksheet_name"], credentials_text, cache_bust
        )
    return source.read_rows()


def scan_source(rows, config):
    email_col = column_to_index(config["email_column"])
    result_col = column_to_index(config["result_column"])

    start_row = int(config["start_row"])
    end_row = min(int(config["end_row"]), len(rows))
    if end_row < start_row:
        raise ValueError("End row must be greater than or equal to start row.")

    pending, skipped, blank, _ = collect_pending(
        rows,
        start_row,
        end_row,
        email_col,
        result_col,
        set(),
    )
    email_to_rows = unique_email_map(pending)

    return {
        "rows": rows,
        "start_row": start_row,
        "end_row": end_row,
        "selected_rows": max(0, end_row - start_row + 1),
        "pending": pending,
        "email_to_rows": email_to_rows,
        "unique_emails": list(email_to_rows),
        "already_done": skipped,
        "blank": blank,
        "email_col": email_col,
        "result_col": result_col,
        "score_col": column_to_index(config["score_column"]),
    }


EMAIL_HEADER_HINTS = ("email", "e-mail", "mail id", "mailid")
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
RESULT_VALUE_HINTS = {"valid", "invalid"}
SCORE_VALUE_PATTERN = re.compile(r"^-?\d+(\.\d+)?(\s*-\s*.+)?$")


def detect_email_column(rows):
    if not rows:
        return None

    header = rows[0]
    for idx, cell in enumerate(header, start=1):
        text = str(cell).strip().lower()
        if any(hint in text for hint in EMAIL_HEADER_HINTS):
            return index_to_column(idx)

    sample = rows[1:200]
    if not sample:
        return None

    max_cols = max((len(row) for row in sample), default=0)
    best_col, best_ratio = None, 0.0
    for col in range(1, max_cols + 1):
        values = [
            str(row[col - 1]).strip()
            for row in sample
            if len(row) >= col and str(row[col - 1]).strip()
        ]
        if len(values) < 3:
            continue
        matches = sum(1 for value in values if EMAIL_PATTERN.match(value))
        ratio = matches / len(values)
        if ratio > best_ratio:
            best_col, best_ratio = col, ratio

    return index_to_column(best_col) if best_ratio >= 0.5 else None


def _column_values(rows, col_idx, sample=300):
    return [
        str(row[col_idx - 1]).strip()
        for row in rows[1:sample]
        if len(row) >= col_idx and str(row[col_idx - 1]).strip()
    ]


def _column_has_foreign_data(values, looks_like_ours):
    if not values:
        return False
    matches = sum(1 for value in values if looks_like_ours(value))
    return (matches / len(values)) < 0.6


def detect_result_score_columns(rows):
    result_idx = column_to_index(DEFAULT_RESULT_COLUMN)
    score_idx = column_to_index(DEFAULT_SCORE_COLUMN)

    result_conflict = _column_has_foreign_data(
        _column_values(rows, result_idx), lambda v: v.lower() in RESULT_VALUE_HINTS
    )
    score_conflict = _column_has_foreign_data(
        _column_values(rows, score_idx), lambda v: bool(SCORE_VALUE_PATTERN.match(v))
    )

    if not result_conflict and not score_conflict:
        return DEFAULT_RESULT_COLUMN, DEFAULT_SCORE_COLUMN

    last_used = max((len(row) for row in rows), default=0)
    shift_start = max(last_used, result_idx, score_idx) + 1
    return index_to_column(shift_start), index_to_column(shift_start + 1)


def _donut_chart(segments, height=230):
    labeled = [(label, value, color) for label, value, color in segments if value]
    fig = go.Figure(
        data=[
            go.Pie(
                labels=[label for label, _, _ in labeled],
                values=[value for _, value, _ in labeled],
                hole=0.65,
                marker=dict(colors=[color for _, _, color in labeled], line=dict(width=0)),
                textinfo="value",
                textposition="inside",
                hovertemplate="%{label}: %{value} (%{percent})<extra></extra>",
                sort=False,
            )
        ]
    )
    fig.update_layout(
        height=height,
        margin=dict(l=10, r=10, t=10, b=10),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5),
    )
    return fig


def _gauge_chart(value, total, height=230):
    pct = 0 if not total else round(min(value / total, 1.0) * 100, 1)
    fig = go.Figure(
        go.Indicator(
            mode="gauge+number",
            value=pct,
            number={"suffix": "%", "font": {"size": 30}},
            gauge={
                "axis": {"range": [0, 100], "tickwidth": 0, "showticklabels": False},
                "bar": {"color": "#2a78d6"},
                "bgcolor": "rgba(128,128,128,.18)",
                "borderwidth": 0,
                "steps": [
                    {"range": [0, 100], "color": "rgba(128,128,128,.08)"},
                ],
            },
        )
    )
    fig.update_layout(height=height, margin=dict(l=20, r=20, t=20, b=10))
    return fig


def render_composition_chart(container, scan, key="composition_chart"):
    with container.container():
        st.caption("Selected rows")
        segments = [
            ("Verified", scan["already_done"], "#2a78d6"),
            ("Pending", len(scan["pending"]), "#f59e0b"),
            ("Blank", scan["blank"], "#94a3b8"),
        ]
        if sum(v for _, v, _ in segments) == 0:
            st.caption("No rows in the selected range.")
            return
        st.plotly_chart(_donut_chart(segments), width="stretch", theme="streamlit", key=key)


def render_outcome_chart(container, valid, invalid, key="outcome_chart"):
    with container.container():
        st.caption("This run: valid vs invalid")
        if valid + invalid == 0:
            st.caption("Run verification to see the breakdown here.")
            return
        segments = [
            ("Valid", valid, "#16a34a"),
            ("Invalid", invalid, "#dc2626"),
        ]
        st.plotly_chart(_donut_chart(segments), width="stretch", theme="streamlit", key=key)


def render_progress_gauge(container, scan, done_this_run, key="gauge_chart"):
    with container.container():
        st.caption("Overall completion")
        total = scan["already_done"] + len(scan["unique_emails"])
        done = scan["already_done"] + done_this_run
        if total == 0:
            st.caption("No rows in the selected range.")
            return
        st.plotly_chart(_gauge_chart(done, total), width="stretch", theme="streamlit", key=key)


def render_kpis(container, scan, done, valid, invalid, total_unique):
    with container.container():
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Rows selected", scan["selected_rows"])
        m2.metric("Emails waiting", len(scan["pending"]))
        m3.metric("Unique emails", total_unique)
        m4.metric("Already done", scan["already_done"])
        m5.metric("Blank emails", scan["blank"])

        r1, r2, r3 = st.columns(3)
        r1.metric("Happened now", done)
        r2.metric("Valid now", valid)
        r3.metric("Invalid now", invalid)

        st.progress(0 if not total_unique else min(done / total_unique, 1.0))


def sidebar_nav():
    st.sidebar.markdown("## EmailVerify Pro")
    st.sidebar.caption("Slow, one-by-one email verification")

    if "page" not in st.session_state:
        st.session_state.page = "Setup"

    pages = ["Setup", "Dashboard", "Quick Verify", "Instructions"]
    page = st.sidebar.radio("Navigation", pages, index=pages.index(st.session_state.page))
    st.session_state.page = page

    config = st.session_state.get("config") or load_saved_config()
    st.sidebar.divider()
    st.sidebar.caption("Current output")
    st.sidebar.write(f"Email: `{config.get('email_column', DEFAULT_EMAIL_COLUMN)}`")
    st.sidebar.write(f"Result: `{config.get('result_column', DEFAULT_RESULT_COLUMN)}`")
    st.sidebar.write(f"Score: `{config.get('score_column', DEFAULT_SCORE_COLUMN)}`")

    return page


def _column_picker(label, header, saved_value, key):
    col_count = max(len(header), 6)
    options = [index_to_column(i) for i in range(1, col_count + 1)]

    def _label(letter, i):
        heading = header[i - 1].strip() if i - 1 < len(header) and header[i - 1] else ""
        return f"{letter} — {heading}" if heading else letter

    labels = [_label(letter, i) for i, letter in enumerate(options, start=1)]
    default_idx = options.index(saved_value) if saved_value in options else 0
    chosen = st.selectbox(label, labels, index=default_idx, key=key)
    return chosen.split(" ")[0]


def _range_and_delay_fields(saved, row_count):
    r1, r2, r3 = st.columns(3)
    start_row = r1.number_input("Start row", min_value=1, value=int(saved.get("start_row", 2)), step=1)
    with r2:
        end_all = st.checkbox("All rows (to end of sheet)", value=bool(saved.get("end_row_all", True)))
        end_row = r2.number_input(
            "End row",
            min_value=1,
            value=int(saved.get("end_row", row_count or 100)),
            step=1,
            disabled=end_all,
        )
    delay_seconds = r3.number_input(
        "Delay after each email", min_value=0, value=int(saved.get("delay_seconds", 8)), step=1
    )
    return start_row, end_all, end_row, delay_seconds


def google_sheets_setup(saved):
    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step 1 · Google Service Account JSON</div>', unsafe_allow_html=True)
        existing_credentials = load_credentials_text()
        saved_client_email = get_json_client_email(existing_credentials)
        if saved_client_email:
            st.success(f"Using saved JSON. Share your sheet with: {saved_client_email}")

        credentials_text = st.text_area(
            "Paste service account JSON",
            value="",
            height=150,
            placeholder="Paste the full JSON here. You can save it on this computer.",
        )
        s1, s2 = st.columns(2)
        remember_json = s1.checkbox("Save this JSON on this computer", value=False)
        if s2.button("Forget saved JSON", width="stretch"):
            remove_saved_credentials()
            st.success("Saved JSON removed.")
            st.rerun()

    active_credentials = credentials_text.strip() or existing_credentials.strip()
    if not active_credentials:
        st.info("Paste a service account JSON (or reuse a saved one) to continue.")
        return None

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step 2 · Spreadsheet URL</div>', unsafe_allow_html=True)
        spreadsheet_url = st.text_input(
            "Paste the Google Sheet URL", saved.get("spreadsheet_url", DEFAULT_SPREADSHEET_URL)
        ).strip()

    if not spreadsheet_url:
        st.info("Paste the spreadsheet URL to continue.")
        return None

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step 3 · Worksheet</div>', unsafe_allow_html=True)
        try:
            titles = cached_worksheet_titles(spreadsheet_url, active_credentials)
        except Exception as exc:
            st.error(f"Could not open that spreadsheet: {exc}")
            return None

        if not titles:
            st.warning("No worksheet tabs found in that spreadsheet.")
            return None

        saved_ws = saved.get("worksheet_name", "")
        default_index = titles.index(saved_ws) if saved_ws in titles else 0
        worksheet_name = st.selectbox("Worksheet / tab", titles, index=default_index)

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step 4 · Columns and Range</div>', unsafe_allow_html=True)
        try:
            rows = cached_google_rows(spreadsheet_url, worksheet_name, active_credentials, 0)
        except Exception as exc:
            st.error(f"Could not read rows: {exc}")
            return None

        header = rows[0] if rows else []
        email_column = detect_email_column(rows)
        if email_column:
            st.success(f"Auto-detected email column: **{email_column}**")
        else:
            st.warning("Could not auto-detect the email column — pick it below.")
            email_column = _column_picker(
                "Email column", header, saved.get("email_column", ""), key="gs_email_col"
            )

        result_column, score_column = detect_result_score_columns(rows)
        st.info(f"Results write to **{result_column}** (Valid/Invalid) and **{score_column}** (score/reason).")

        start_row, end_all, end_row, delay_seconds = _range_and_delay_fields(saved, len(rows))

    config = {
        "mode": "Google Sheets",
        "spreadsheet_url": spreadsheet_url,
        "worksheet_name": worksheet_name,
        "excel_sheet": "",
        "email_column": email_column,
        "result_column": result_column,
        "score_column": score_column,
        "start_row": int(start_row),
        "end_row": 10**9 if end_all else int(end_row),
        "end_row_all": end_all,
        "delay_seconds": int(delay_seconds),
        "retry_delay": 30,
        "jitter": 3,
    }
    return config, credentials_text, remember_json


def excel_setup(saved):
    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Excel Upload</div>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader("Upload .xlsx file", type=["xlsx"])
        if uploaded_file:
            st.session_state.excel_file = uploaded_file

    uploaded_file = st.session_state.get("excel_file")
    if uploaded_file is None:
        st.info("Upload an Excel file to continue.")
        return None

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step · Worksheet</div>', unsafe_allow_html=True)
        try:
            from openpyxl import load_workbook

            titles = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True).sheetnames
        except Exception as exc:
            st.error(f"Could not read workbook: {exc}")
            return None

        saved_sheet = saved.get("excel_sheet", "")
        default_index = titles.index(saved_sheet) if saved_sheet in titles else 0
        excel_sheet = st.selectbox("Worksheet", titles, index=default_index)

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Step · Columns and Range</div>', unsafe_allow_html=True)
        try:
            rows = read_excel_rows_from_bytes(uploaded_file.getvalue(), excel_sheet)
        except Exception as exc:
            st.error(f"Could not read rows: {exc}")
            return None

        header = rows[0] if rows else []
        email_column = detect_email_column(rows)
        if email_column:
            st.success(f"Auto-detected email column: **{email_column}**")
        else:
            st.warning("Could not auto-detect the email column — pick it below.")
            email_column = _column_picker(
                "Email column", header, saved.get("email_column", ""), key="xl_email_col"
            )

        result_column, score_column = detect_result_score_columns(rows)
        st.info(f"Results write to **{result_column}** (Valid/Invalid) and **{score_column}** (score/reason).")

        start_row, end_all, end_row, delay_seconds = _range_and_delay_fields(saved, len(rows))

    config = {
        "mode": "Excel",
        "spreadsheet_url": "",
        "worksheet_name": "",
        "excel_sheet": excel_sheet,
        "email_column": email_column,
        "result_column": result_column,
        "score_column": score_column,
        "start_row": int(start_row),
        "end_row": 10**9 if end_all else int(end_row),
        "end_row_all": end_all,
        "delay_seconds": int(delay_seconds),
        "retry_delay": 30,
        "jitter": 3,
    }
    return config, "", False


def _persist_setup(config, credentials_text="", remember_json=False):
    if config["mode"] == "Google Sheets" and credentials_text.strip() and remember_json:
        client_email = save_credentials(credentials_text)
        st.success(f"Google JSON saved. Share your sheet with: {client_email}")

    if config["mode"] == "Google Sheets" and credentials_text.strip():
        config = {**config, "credentials_text": credentials_text.strip()}

    save_config({k: v for k, v in config.items() if k != "credentials_text"})
    st.session_state.config = config


def setup_page():
    app_header("Setup", "Connect a Google Sheet or upload Excel — most of it fills in automatically.")
    saved = load_saved_config()

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Data Source</div>', unsafe_allow_html=True)
        mode = st.segmented_control(
            "Source",
            ["Google Sheets", "Excel"],
            default=saved.get("mode", "Google Sheets"),
            label_visibility="collapsed",
        )

        batch_col, coming_col = st.columns([1, 2])
        batch_col.toggle("Verify 1 by 1", value=True, disabled=True)
        coming_col.button("Batch mode coming soon", disabled=True, width="stretch")

    result = google_sheets_setup(saved) if mode == "Google Sheets" else excel_setup(saved)
    if result is None:
        return
    config, credentials_text, remember_json = result

    c1, c2 = st.columns([1, 1])
    if c1.button("Save setup", type="primary", width="stretch"):
        _persist_setup(config, credentials_text, remember_json)
        st.success("Setup saved.")

    if c2.button("Save and open dashboard", width="stretch"):
        _persist_setup(config, credentials_text, remember_json)
        st.session_state.page = "Dashboard"
        st.rerun()


def dashboard_page():
    top_col, refresh_col = st.columns([5, 1])
    with top_col:
        app_header("Dashboard", "Review volume, then verify emails one at a time.")
    with refresh_col:
        if st.button("Refresh sheet data", width="stretch"):
            st.session_state.sheet_cache_bust = st.session_state.get("sheet_cache_bust", 0) + 1
            st.rerun()

    config = st.session_state.get("config") or load_saved_config()
    if not config:
        st.warning("Complete Setup first.")
        if st.button("Go to Setup"):
            st.session_state.page = "Setup"
            st.rerun()
        return

    try:
        source, excel_tmp_path = make_source(config)
        rows = read_rows_for(source, config)
        scan = scan_source(rows, config)
    except Exception as exc:
        st.error(str(exc))
        st.stop()

    total_unique = len(scan["unique_emails"])
    done_this_run = st.session_state.get("done_this_run", 0)
    valid_this_run = st.session_state.get("valid_this_run", 0)
    invalid_this_run = st.session_state.get("invalid_this_run", 0)

    kpi_slot = st.empty()
    render_kpis(kpi_slot, scan, done_this_run, valid_this_run, invalid_this_run, total_unique)

    with st.container(border=True):
        chart_col1, chart_col2, chart_col3 = st.columns(3)
        with chart_col1:
            composition_slot = st.empty()
            render_composition_chart(composition_slot, scan, key="composition_chart_initial")
        with chart_col2:
            gauge_slot = st.empty()
            render_progress_gauge(gauge_slot, scan, done_this_run, key="gauge_chart_initial")
        with chart_col3:
            outcome_slot = st.empty()
            render_outcome_chart(
                outcome_slot, valid_this_run, invalid_this_run, key="outcome_chart_initial"
            )

    start = st.button(
        "Start verifying",
        type="primary",
        width="stretch",
        disabled=total_unique == 0,
    )

    status_box = st.empty()
    log_box = st.empty()
    run_progress = st.empty()

    if start:
        logs = []
        verified = 0
        total_to_run = total_unique
        for email in scan["unique_emails"]:
            row_numbers = scan["email_to_rows"][email]
            status_box.info(f"Verifying {email}")

            try:
                results = verify_batch([email], config.get("retry_delay", 30), config.get("jitter", 3))
                result = results[0] if results else {"email": email, "status": "FAILED"}
                verification, score_reason = convert_result(result)

                row_updates = [
                    (row_number, verification, score_reason)
                    for row_number in row_numbers
                ]
                source.write_results(row_updates, scan["result_col"], scan["score_col"])
                source.save()

                done_this_run += len(row_updates)
                if verification == "Valid":
                    valid_this_run += len(row_updates)
                else:
                    invalid_this_run += len(row_updates)

                st.session_state.done_this_run = done_this_run
                st.session_state.valid_this_run = valid_this_run
                st.session_state.invalid_this_run = invalid_this_run

                rows_text = ", ".join(str(row) for row in row_numbers)
                logs.append(f"Rows {rows_text}: {email} -> {verification} | {score_reason}")
            except Exception as exc:
                logs.append(f"{email} -> ERROR | {exc}")

            verified += 1
            log_box.code("\n".join(logs[-80:]), language="text")
            run_progress.progress(min(verified / total_to_run, 1.0))
            render_kpis(kpi_slot, scan, done_this_run, valid_this_run, invalid_this_run, total_unique)
            render_outcome_chart(
                outcome_slot, valid_this_run, invalid_this_run, key=f"outcome_chart_run_{verified}"
            )
            render_progress_gauge(gauge_slot, scan, done_this_run, key=f"gauge_chart_run_{verified}")

            if verified < total_to_run:
                delay = int(config.get("delay_seconds", 8))
                status_box.info(f"Saved. Waiting {delay}s before next email...")
                time.sleep(delay)

        status_box.success(f"Finished {verified} unique emails.")
        st.session_state.sheet_cache_bust = st.session_state.get("sheet_cache_bust", 0) + 1

        if config["mode"] == "Excel":
            output = BytesIO()
            source.workbook.save(output)
            st.download_button(
                "Download verified Excel",
                data=output.getvalue(),
                file_name="verified_emails.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    if config["mode"] == "Excel" and excel_tmp_path:
        try:
            os.remove(excel_tmp_path)
        except OSError:
            pass


MAX_QUICK_BULK = 500


def results_to_csv_bytes(rows):
    return pd.DataFrame(rows).to_csv(index=False).encode("utf-8")


def results_to_excel_bytes(rows):
    buffer = BytesIO()
    pd.DataFrame(rows).to_excel(buffer, index=False, sheet_name="Results")
    return buffer.getvalue()


def render_download_buttons(rows, key_prefix):
    d1, d2 = st.columns(2)
    d1.download_button(
        "Download CSV",
        data=results_to_csv_bytes(rows),
        file_name="verified_emails.csv",
        mime="text/csv",
        width="stretch",
        key=f"{key_prefix}_csv",
    )
    d2.download_button(
        "Download Excel",
        data=results_to_excel_bytes(rows),
        file_name="verified_emails.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        key=f"{key_prefix}_xlsx",
    )


def quick_verify_page():
    app_header("Quick Verify", "Check one email, or paste up to 500 to verify and download.")

    tab_single, tab_bulk = st.tabs(["Single email", "Bulk paste"])

    with tab_single:
        with st.container(border=True):
            email_input = st.text_input("Email address", placeholder="someone@example.com")
            if st.button("Verify", type="primary", key="single_verify_btn"):
                email = clean_email(email_input)
                if not email:
                    st.warning("Enter an email address.")
                else:
                    with st.spinner(f"Verifying {email}..."):
                        try:
                            results = verify_batch([email], 30, 3)
                            result = results[0] if results else {"email": email, "status": "FAILED"}
                            verification, score_reason = convert_result(result)
                        except Exception as exc:
                            verification, score_reason = None, str(exc)

                    if verification == "Valid":
                        st.success(f"{email} -> Valid | {score_reason}")
                    elif verification == "Invalid":
                        st.error(f"{email} -> Invalid | {score_reason}")
                    else:
                        st.error(f"Could not verify {email}: {score_reason}")

    with tab_bulk:
        with st.container(border=True):
            pasted = st.text_area(
                "Paste emails — one per line, or separated by commas/spaces",
                height=220,
                placeholder="jane@example.com\njohn@example.com\n...",
            )
            emails = parse_pasted_emails(pasted)[:MAX_QUICK_BULK]
            total_found = len(parse_pasted_emails(pasted))

            caption = f"{len(emails)} unique email(s) ready to verify"
            if total_found > MAX_QUICK_BULK:
                caption += f" (only the first {MAX_QUICK_BULK} of {total_found} pasted will be verified)"
            st.caption(caption)

            start_bulk = st.button(
                "Verify all", type="primary", disabled=not emails, key="bulk_verify_btn"
            )

        if start_bulk:
            progress = st.progress(0.0)
            status = st.empty()
            rows = []
            chunk_size = 50

            for i in range(0, len(emails), chunk_size):
                chunk = emails[i : i + chunk_size]
                status.info(f"Verifying {i + 1}-{i + len(chunk)} of {len(emails)}...")

                try:
                    results = verify_batch(chunk, 30, 3)
                except Exception:
                    results = []
                result_by_email = {clean_email(r.get("email", "")): r for r in results}

                for email in chunk:
                    result = result_by_email.get(email, {"email": email, "status": "FAILED"})
                    verification, score_reason = convert_result(result)
                    rows.append({"Email": email, "Result": verification, "Score / Reason": score_reason})

                progress.progress(min((i + len(chunk)) / len(emails), 1.0))
                if i + chunk_size < len(emails):
                    time.sleep(2)

            status.success(f"Verified {len(rows)} email(s).")
            st.session_state.quick_verify_results = rows

        results_rows = st.session_state.get("quick_verify_results")
        if results_rows:
            valid_count = sum(1 for row in results_rows if row["Result"] == "Valid")

            c1, c2, c3 = st.columns(3)
            c1.metric("Total", len(results_rows))
            c2.metric("Valid", valid_count)
            c3.metric("Invalid", len(results_rows) - valid_count)

            st.dataframe(pd.DataFrame(results_rows), width="stretch", hide_index=True)
            render_download_buttons(results_rows, key_prefix="bulk")


def instructions_page():
    app_header("Instructions", "A new-user setup guide with direct credential links.")

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Columns</div>', unsafe_allow_html=True)
        st.markdown(
            """
| Column | Use |
| --- | --- |
| C | Email address |
| W | Valid / Invalid result |
| X | Score and reason |
"""
        )

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Direct Links</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        c1.link_button("Create Google Project", "https://console.cloud.google.com/projectcreate", width="stretch")
        c2.link_button("Enable Sheets API", "https://console.cloud.google.com/apis/library/sheets.googleapis.com", width="stretch")
        c3.link_button("Enable Drive API", "https://console.cloud.google.com/apis/library/drive.googleapis.com", width="stretch")
        c4, c5, c6 = st.columns(3)
        c4.link_button("Create Service Account", "https://console.cloud.google.com/iam-admin/serviceaccounts/create", width="stretch")
        c5.link_button("Service Accounts", "https://console.cloud.google.com/iam-admin/serviceaccounts", width="stretch")
        c6.link_button("Open Google Sheets", "https://sheets.google.com", width="stretch")

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Get Google JSON</div>', unsafe_allow_html=True)
        st.markdown(
            """
<div class="ev-step">1. Create/select a Google Cloud project.</div>
<div class="ev-step">2. Enable Google Sheets API and Google Drive API.</div>
<div class="ev-step">3. Create a service account named <code>email-verify</code>.</div>
<div class="ev-step">4. Open that service account, go to <b>Keys</b>, click <b>Add key</b>, choose <b>Create new key</b>, then select <b>JSON</b>.</div>
<div class="ev-step">5. Open the downloaded JSON and copy the whole file into the Setup page.</div>
""",
            unsafe_allow_html=True,
        )
        st.warning("Google lets you download the private JSON key only once. Keep it private.")

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">Share Your Sheet With The JSON Email</div>', unsafe_allow_html=True)
        st.markdown(
            """
1. Open your downloaded JSON file.
2. Find this field:
"""
        )
        st.code('"client_email": "email-verify@your-project.iam.gserviceaccount.com"', language="json")
        st.markdown(
            """
3. Copy only the email address inside `client_email`.
4. Open your Google Sheet.
5. Click **Share**.
6. Paste that `client_email`.
7. Set permission to **Editor**.
8. Click **Send**.
"""
        )

    with st.container(border=True):
        st.markdown('<div class="ev-kicker">First Safe Run</div>', unsafe_allow_html=True)
        st.code(
            """Source: Google Sheets
Email column: C
Result column: W
Score / reason column: X
Start row: 2
End row: 10
Delay after each email: 8 to 15 seconds""",
            language="text",
        )


def main():
    inject_css()
    page = sidebar_nav()

    if page == "Setup":
        setup_page()
    elif page == "Dashboard":
        dashboard_page()
    elif page == "Quick Verify":
        quick_verify_page()
    else:
        instructions_page()


if __name__ == "__main__":
    main()
